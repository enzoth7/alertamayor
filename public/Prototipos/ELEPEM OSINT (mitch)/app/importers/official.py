import csv
import hashlib
import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from typing import Any
from urllib.parse import urlsplit

from geoalchemy2.elements import WKTElement
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.audit import append_audit_event
from app.enums import ImportFormat, OfficialStatus
from app.models import OfficialRegistryRecord, RegistryImport
from app.normalization import normalize_fields, normalize_text

_ALIASES = {
    "identifier": ("id", "identificador", "numero", "nro", "registro", "numero registro"),
    "name": (
        "nombre",
        "denominacion",
        "establecimiento",
        "nombre establecimiento",
        "razon social",
    ),
    "status": ("estado", "situacion", "habilitacion", "estado habilitacion"),
    "address": ("direccion", "domicilio", "direccion establecimiento"),
    "department": ("departamento", "depto"),
    "phone": ("telefono", "telefono institucional", "telefono establecimiento"),
    "website": ("sitio web", "web", "website", "url"),
    "latitude": ("latitud", "latitude", "lat"),
    "longitude": ("longitud", "longitude", "lon", "lng"),
}


class RegistryImportError(ValueError):
    pass


@dataclass(frozen=True, slots=True)
class ParsedOfficialRecord:
    row_number: int
    official_identifier: str | None
    business_name: str
    official_status: OfficialStatus
    address: str | None
    department: str | None
    phone: str | None
    website: str | None
    latitude: float | None
    longitude: float | None
    source_columns: tuple[str, ...]


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    return str(value)


def _row_to_string_dict(row: dict[Any, Any]) -> dict[str, Any]:
    return {
        str(key).strip(): _json_safe(value)
        for key, value in row.items()
        if key is not None and str(key).strip()
    }


def _parse_csv(data: bytes) -> list[dict[str, Any]]:
    try:
        text = data.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise RegistryImportError("CSV must be UTF-8 encoded") from exc
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [
        _row_to_string_dict(dict(row)) for row in csv.DictReader(io.StringIO(text), dialect=dialect)
    ]


def _parse_json(data: bytes) -> list[dict[str, Any]]:
    try:
        decoded = json.loads(data.decode("utf-8-sig"))
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise RegistryImportError("Invalid UTF-8 JSON") from exc
    if isinstance(decoded, dict):
        for key in ("records", "data", "results", "establecimientos"):
            if isinstance(decoded.get(key), list):
                decoded = decoded[key]
                break
    if not isinstance(decoded, list) or not all(isinstance(item, dict) for item in decoded):
        raise RegistryImportError(
            "JSON must be an array of records or contain records/data/results"
        )
    return [_row_to_string_dict(item) for item in decoded]


def _parse_xlsx(data: bytes) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise RegistryImportError("Invalid XLSX workbook") from exc
    worksheet = workbook.active
    if worksheet is None:
        workbook.close()
        raise RegistryImportError("XLSX workbook has no active worksheet")
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    except StopIteration:
        return []
    result = []
    for values in rows:
        row = {header: value for header, value in zip(headers, values, strict=False) if header}
        if any(value not in (None, "") for value in row.values()):
            result.append(_row_to_string_dict(row))
    workbook.close()
    return result


def parse_structured_rows(data: bytes, source_format: ImportFormat) -> list[dict[str, Any]]:
    if not data:
        raise RegistryImportError("Import file is empty")
    if source_format in (ImportFormat.CSV, ImportFormat.PDF_DERIVED_CSV):
        return _parse_csv(data)
    if source_format in (ImportFormat.JSON, ImportFormat.PDF_DERIVED_JSON):
        return _parse_json(data)
    if source_format is ImportFormat.XLSX:
        return _parse_xlsx(data)
    raise RegistryImportError(f"Unsupported import format: {source_format}")


def _normalized_row(row: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in row.items():
        normalized_key = normalize_text(key)
        if normalized_key:
            result[normalized_key] = value
    return result


def _find_value(row: dict[str, Any], field: str) -> Any:
    for alias in _ALIASES[field]:
        if alias in row and row[alias] not in (None, ""):
            return row[alias]
    return None


def _optional_string(value: Any) -> str | None:
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def _float_value(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(str(value).strip().replace(",", "."))
    except ValueError:
        return None


def _coordinates(latitude: Any, longitude: Any) -> tuple[float | None, float | None]:
    lat = _float_value(latitude)
    lon = _float_value(longitude)
    if lat is None or lon is None:
        return None, None
    if -35.5 <= lat <= -30.0 and -59.0 <= lon <= -52.5:
        return lat, lon
    return None, None


def parse_official_status(value: Any) -> OfficialStatus:
    normalized = normalize_text(_optional_string(value))
    if normalized is None:
        return OfficialStatus.UNKNOWN
    if any(token in normalized for token in ("tramite", "proceso", "pendiente")):
        return OfficialStatus.IN_PROCESS
    if any(
        token in normalized
        for token in ("no habilitad", "inhabilitad", "suspendid", "vencid", "baja")
    ):
        return OfficialStatus.OTHER
    if any(token in normalized for token in ("habilitad", "autorizad", "vigente")):
        return OfficialStatus.HABILITATED
    return OfficialStatus.OTHER


def map_official_rows(rows: list[dict[str, Any]]) -> list[ParsedOfficialRecord]:
    mapped: list[ParsedOfficialRecord] = []
    for row_number, raw_row in enumerate(rows, start=2):
        row = _normalized_row(raw_row)
        name = _optional_string(_find_value(row, "name"))
        if name is None:
            continue
        latitude, longitude = _coordinates(
            _find_value(row, "latitude"), _find_value(row, "longitude")
        )
        mapped.append(
            ParsedOfficialRecord(
                row_number=row_number,
                official_identifier=_optional_string(_find_value(row, "identifier")),
                business_name=name,
                official_status=parse_official_status(_find_value(row, "status")),
                address=_optional_string(_find_value(row, "address")),
                department=_optional_string(_find_value(row, "department")),
                phone=_optional_string(_find_value(row, "phone")),
                website=_optional_string(_find_value(row, "website")),
                latitude=latitude,
                longitude=longitude,
                source_columns=tuple(sorted(raw_row)),
            )
        )
    if not mapped:
        raise RegistryImportError("No rows with a recognized facility name column were found")
    return mapped


def import_official_registry(
    session: Session,
    *,
    data: bytes,
    source_filename: str,
    source_format: ImportFormat,
    source_url: str,
    retrieved_at: datetime,
    actor: str,
) -> RegistryImport:
    if retrieved_at.tzinfo is None or retrieved_at.utcoffset() is None:
        raise RegistryImportError("retrieved_at must include a timezone offset")
    parsed_source_url = urlsplit(source_url)
    if parsed_source_url.scheme not in {"http", "https"} or not parsed_source_url.hostname:
        raise RegistryImportError("source_url must be HTTP(S)")
    source_host = parsed_source_url.hostname.lower()
    if source_host != "gub.uy" and not source_host.endswith(".gub.uy"):
        raise RegistryImportError("official registry source_url must be on a gub.uy domain")
    rows = parse_structured_rows(data, source_format)
    mapped_rows = map_official_rows(rows)
    registry_import = RegistryImport(
        source_url=source_url,
        source_filename=source_filename,
        source_format=source_format,
        retrieved_at=retrieved_at,
        sha256=hashlib.sha256(data).hexdigest(),
        row_count=len(mapped_rows),
    )
    session.add(registry_import)
    session.flush()
    for item in mapped_rows:
        normalized = normalize_fields(
            name=item.business_name,
            address=item.address,
            phone=item.phone,
            website=item.website,
        )
        location = (
            WKTElement(f"POINT({item.longitude} {item.latitude})", srid=4326)
            if item.latitude is not None and item.longitude is not None
            else None
        )
        session.add(
            OfficialRegistryRecord(
                registry_import_id=registry_import.id,
                official_identifier=item.official_identifier,
                business_name=item.business_name,
                normalized_name=normalized.name or "",
                official_status=item.official_status,
                address=item.address,
                normalized_address=normalized.address,
                department=item.department,
                phone=item.phone,
                normalized_phone=normalized.phone,
                website=item.website,
                normalized_domain=normalized.domain,
                latitude=item.latitude,
                longitude=item.longitude,
                location=location,
                source_row={
                    "row_number": item.row_number,
                    "source_columns": item.source_columns,
                },
            )
        )
    append_audit_event(
        session,
        event_type="OFFICIAL_REGISTRY_IMPORTED",
        actor=actor,
        entity_type="registry_import",
        entity_id=registry_import.id,
        payload={
            "source_url": source_url,
            "retrieved_at": retrieved_at.isoformat(),
            "sha256": registry_import.sha256,
            "source_format": source_format.value,
            "row_count": len(mapped_rows),
        },
    )
    session.commit()
    return registry_import
